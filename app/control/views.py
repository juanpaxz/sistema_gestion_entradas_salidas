from django.shortcuts import render
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import Group
from django.utils import timezone
from django.contrib.auth.views import LoginView
from django.db import IntegrityError
from datetime import datetime, date
import datetime as _dt
import logging
from .models import Empleado, Asistencia, Horario, Justificante, SystemConfig
from .forms import EmpleadoCreationForm, EmpleadoForm, JustificanteRetardoForm, HorarioForm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# Configurar logger para la aplicación
logger = logging.getLogger(__name__)

class CustomLoginView(LoginView):
    template_name = 'control/login.html'
    
    def get_success_url(self):
        user = self.request.user
        
        if user.groups.filter(name='administracion').exists():
            return '/control/admin/dashboard/'
        elif user.groups.filter(name='empleado').exists():
            return '/control/empleado/dashboard/'
        elif user.groups.filter(name='supervisores').exists():
            return '/supervisores/panel/'
        else:
            return '/default/'  # Página por defecto

def es_administracion(user):
    return user.groups.filter(name='administracion').exists()

# Create your views here.
def home(request):
    """Vista principal de la app `control` para verificar que la app responde."""
    return HttpResponse("Control app: funciona correctamente.")

@login_required
def listar_empleados(request):
    """Lista los empleados. Acceso solo para administradores."""
    user = request.user
    if not es_administracion(user):
        return HttpResponseForbidden('No tienes permiso para ver esta página')

    empleados = Empleado.objects.select_related('user').all().order_by('nombre', 'apellido')
    return render(request, 'control/administracion/listar_empleados.html', {'empleados': empleados})

@login_required
def dashboard(request):
    """Dashboard para administradores.

    Muestra métricas básicas del sistema (número de empleados, activos) y
    solo está disponible para usuarios del grupo 'administracion'.
    """
    user = request.user
    # comprobar pertenencia al grupo administracion
    if not es_administracion(user):
        return HttpResponseForbidden('No tienes permiso para ver esta página')

    # Si se envía un formulario para actualizar la configuración (umbral de retardo)
    if request.method == 'POST':
        ret_min = request.POST.get('retardo_minutos')
        if ret_min is not None:
            try:
                val = int(ret_min)
                cfg = SystemConfig.get_solo()
                cfg.retardo_minutos = max(0, val)
                cfg.save()
                messages.success(request, f'Umbral de retardo actualizado a {cfg.retardo_minutos} minutos.')
                return redirect('control:admin_dashboard')
            except ValueError:
                messages.error(request, 'Valor inválido para minutos de retardo')

    total_empleados = Empleado.objects.count()
    activos = Empleado.objects.filter(estado='activo').count()

    # 🔥 NUEVO: obtener empleados sin horario asignado
    empleados_sin_horario = Empleado.objects.filter(horarios__isnull=True)

    # Obtener umbral actual para mostrar en el dashboard
    try:
        cfg = SystemConfig.get_solo()
        retardo_actual = cfg.retardo_minutos
    except Exception:
        retardo_actual = 0

    context = {
        'total_empleados': total_empleados,
        'empleados_activos': activos,
        'retardo_minutos': retardo_actual,

        # 🔥 NUEVO: enviar la lista al template
        'empleados_sin_horario': empleados_sin_horario,
    }
    return render(request, 'control/administracion/dashboard.html', context)

@login_required
def crear_empleado(request):
    """Crear un nuevo empleado (crea también el usuario asociado).

    Solo usuarios del grupo 'administracion' pueden acceder a esta vista.
    """
    user = request.user
    if not es_administracion(user):
        return HttpResponseForbidden('No tienes permiso para ver esta página')

    if request.method == 'POST':
        form = EmpleadoCreationForm(request.POST)
        if form.is_valid():
            # Guardar usuario y empleado
            new_user = form.save(commit=True)
            logger.info(f"Usuario creado exitosamente: {new_user.username}")
            
            # Crear registro Empleado y manejar errores de integridad (RFC único)
            try:
                empleado = Empleado.objects.create(
                    user=new_user,
                    nombre=form.cleaned_data.get('nombre'),
                    apellido=form.cleaned_data.get('apellido'),
                    puesto=form.cleaned_data.get('puesto'),
                    estado=form.cleaned_data.get('estado'),
                    rfc=form.cleaned_data.get('rfc'),
                    huella_biometrica=form.cleaned_data.get('huella_biometrica') or None,
                )
                # Asignar horarios opcionales seleccionados en el formulario
                horarios_selected = form.cleaned_data.get('horarios')
                if horarios_selected:
                    empleado.horarios.set(horarios_selected)
                logger.info(f"Empleado creado exitosamente: {empleado.nombre} {empleado.apellido} (RFC: {empleado.rfc})")
                
                # Asignar al grupo seleccionado en el formulario (role). Fallback a 'empleado'
                selected_role = form.cleaned_data.get('role', 'empleado')
                try:
                    grupo = Group.objects.get(name=selected_role)
                    new_user.groups.add(grupo)
                    logger.info(f"Usuario {new_user.username} añadido al grupo '{selected_role}'")
                except Group.DoesNotExist:
                    logger.warning(f"El grupo '{selected_role}' no existe en el sistema")
                
                # Si todo fue exitoso, mostrar mensaje y redirigir
                messages.success(request, f'Empleado {empleado.nombre} {empleado.apellido} creado correctamente')
                return redirect('control:admin_dashboard')
                    
            except IntegrityError as e:
                logger.error(f"Error de integridad al crear empleado con RFC {form.cleaned_data.get('rfc')}: {str(e)}")
                new_user.delete()
                form.add_error('rfc', 'Ya existe un empleado con ese RFC o ocurrió un conflicto en la base de datos.')
                
            except Exception as e:
                logger.error(f"Error inesperado al crear empleado: {str(e)}", exc_info=True)
                new_user.delete()
                messages.error(request, f'Ocurrió un error al crear el empleado: {str(e)}')

                messages.success(request, 'Empleado creado correctamente.')
                return redirect('control:admin_dashboard')
    else:
        form = EmpleadoCreationForm()

    return render(request, 'control/administracion/crear_empleado.html', {'form': form})


@login_required
def editar_empleado(request, empleado_id):
    """Editar los datos de un empleado existente."""
    user = request.user
    if not es_administracion(user):
        return HttpResponseForbidden('No tienes permiso para ver esta página')

    empleado = get_object_or_404(Empleado, pk=empleado_id)
    if request.method == 'POST':
        form = EmpleadoForm(request.POST, instance=empleado)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empleado actualizado correctamente.')
            return redirect('control:listar')
    else:
        form = EmpleadoForm(instance=empleado)

    return render(request, 'control/administracion/editar_empleado.html', {'form': form, 'empleado': empleado})


@login_required
def eliminar_empleado(request, empleado_id):
    """Eliminar un empleado (confirma antes de borrar)."""
    user = request.user
    if not es_administracion(user):
        return HttpResponseForbidden('No tienes permiso para ver esta página')

    empleado = get_object_or_404(Empleado, pk=empleado_id)
    if request.method == 'POST':
        # delete associated user (this cascades to empleado normally)
        try:
            empleado.user.delete()
        except Exception:
            empleado.delete()
        messages.success(request, 'Empleado eliminado correctamente.')
        return redirect('control:listar')

    return render(request, 'control/administracion/confirm_delete_empleado.html', {'empleado': empleado})


@login_required
def listar_horarios(request):
    """Lista los horarios -- acceso solo administradores."""
    user = request.user
    if not es_administracion(user):
        return HttpResponseForbidden('No tienes permiso para ver esta página')

    horarios = Horario.objects.all().order_by('nombre')
    return render(request, 'control/administracion/horarios_list.html', {'horarios': horarios})


@login_required
def crear_horario(request):
    user = request.user
    if not es_administracion(user):
        return HttpResponseForbidden('No tienes permiso para ver esta página')

    if request.method == 'POST':
        form = HorarioForm(request.POST)
        if form.is_valid():
            horario = form.save()
            messages.success(request, 'Horario creado correctamente.')
            return redirect('control:listar_horarios')
    else:
        form = HorarioForm()

    return render(request, 'control/administracion/crear_horario.html', {'form': form})


@login_required
def editar_horario(request, horario_id):
    user = request.user
    if not es_administracion(user):
        return HttpResponseForbidden('No tienes permiso para ver esta página')

    horario = get_object_or_404(Horario, pk=horario_id)
    if request.method == 'POST':
        form = HorarioForm(request.POST, instance=horario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Horario actualizado correctamente.')
            return redirect('control:listar_horarios')
    else:
        form = HorarioForm(instance=horario)

    return render(request, 'control/administracion/crear_horario.html', {'form': form, 'horario': horario})


@login_required
def eliminar_horario(request, horario_id):
    user = request.user
    if not es_administracion(user):
        return HttpResponseForbidden('No tienes permiso para ver esta página')

    horario = get_object_or_404(Horario, pk=horario_id)
    if request.method == 'POST':
        horario.delete()
        messages.success(request, 'Horario eliminado correctamente.')
        return redirect('control:listar_horarios')

    return render(request, 'control/administracion/confirm_delete_horario.html', {'horario': horario})


def empleados_sin_horario(request):
    empleados_con_horario = Horario.objects.values_list('empleado_id', flat=True)
    empleados_sin = Empleado.objects.exclude(id__in=empleados_con_horario)

    return render(request, 'empleados_sin_horario.html', {
        'empleados': empleados_sin
    })


def registro_asistencia(request):
    """Vista completamente pública para el registro de asistencias."""

    
    return render(request, 'control/asistencias/registro.html')


def registrar_entrada(request):
    """Registrar la entrada de un empleado."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

    # Priorizar RFC enviado en el POST (permitir kioscos/public access)
    rfc = request.POST.get('rfc')
    empleado = None

    if rfc:
        try:
            empleado = Empleado.objects.get(rfc__iexact=rfc)
        except Empleado.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'No se encontró empleado con ese RFC'}, status=404)

    # Verificar si ya existe una asistencia para hoy
    asistencia, created = Asistencia.objects.get_or_create(
        empleado=empleado,
        fecha=date.today()
    )
    if asistencia.hora_entrada:
        return JsonResponse({
            'status': 'error',
            'message': 'Ya has registrado tu entrada hoy'
        })

    # Registrar la hora de entrada
    asistencia.registrar_entrada()

    # Verificar si es un retardo en base al horario asignado al empleado.
    # `hora_entrada` puede ser `time` (TimeField) o `datetime`.
    # Calcular minutos de diferencia usando el helper del modelo
    try:
        mins = asistencia.compute_diferencia_minutes()
    except Exception:
        mins = None

    # Obtener umbral de retardo desde configuración
    try:
        cfg = SystemConfig.get_solo()
        umbral = int(cfg.retardo_minutos or 0)
    except Exception:
        umbral = 0
    
    print("DEBUG >>> hora_entrada:", asistencia.hora_entrada)
    print("DEBUG >>> mins calculados:", mins)
    print("DEBUG >>> tipo mins:", type(mins))
    print("DEBUG >>> umbral minutos:", umbral)
    print("DEBUG >>> ENTRO AL IF?:", mins is not None and mins > umbral)

    # Si la diferencia supera el umbral, marcar retardo
    if mins is not None and mins > umbral:
        asistencia.tipo = 'retardo'
        asistencia.save()

    # Formatear hora para la respuesta (hora almacenada es TimeField)
    hora_str = None
    if isinstance(asistencia.hora_entrada, _dt.time):
        hora_str = asistencia.hora_entrada.strftime('%H:%M:%S')

    return JsonResponse({
        'status': 'success',
        'message': 'Entrada registrada exitosamente',
        'hora': hora_str,
        'diferencia_minutos': mins,
        'umbral_minutos': umbral,
    })


def registrar_salida(request):
    """Registrar la salida de un empleado."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

    # Priorizar RFC enviado en el POST
    rfc = request.POST.get('rfc')
    empleado = None

    if rfc:
        try:
            empleado = Empleado.objects.get(rfc__iexact=rfc)
        except Empleado.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'No se encontró empleado con ese RFC'}, status=404)

    # Buscar la asistencia de hoy
    try:
        asistencia = Asistencia.objects.get(
            empleado=empleado,
            fecha=date.today()
        )

        if not asistencia.hora_entrada:
            return JsonResponse({
                'status': 'error',
                'message': 'Debes registrar primero tu entrada'
            })

        if asistencia.hora_salida:
            return JsonResponse({
                'status': 'error',
                'message': 'Ya has registrado tu salida hoy'
            })

        # Registrar la hora de salida
        asistencia.registrar_salida()

        # Obtener hora_salida como time
        hora_salida = None
        if asistencia.hora_salida:
            if isinstance(asistencia.hora_salida, _dt.time):
                hora_salida = asistencia.hora_salida
            else:
                try:
                    hora_salida = timezone.localtime(asistencia.hora_salida).time()
                except Exception:
                    hora_salida = None

        return JsonResponse({
            'status': 'success',
            'message': 'Salida registrada exitosamente',
            # Mostrar la hora en la zona local del servidor
            'hora': hora_salida.strftime('%H:%M:%S') if hora_salida else None
        })

    except Asistencia.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'No se encontró registro de entrada para hoy'
        })


@login_required
def ver_asistencias(request):
    """Ver el historial de asistencias."""
    user = request.user

    # Si es administrador, puede ver todas las asistencias
    if es_administracion(user):
        asistencias = Asistencia.objects.all().select_related('empleado')
    else:
        # Si es empleado, solo ve sus propias asistencias
        try:
            empleado = Empleado.objects.get(user=user)
            asistencias = Asistencia.objects.filter(empleado=empleado)
            # Obtener el horario aplicable para la fecha actual
            horario_para_hoy = empleado.get_horario_para_fecha()
        except Empleado.DoesNotExist:
            return HttpResponseForbidden('No tienes acceso a esta página')

    # Ordenar por fecha descendente
    asistencias = asistencias.order_by('-fecha', '-hora_entrada')

    context = {
        'asistencias': asistencias,
    }
    # Si se resolvió un empleado, incluir horarios en el contexto para la plantilla
    if not es_administracion(user):
        context.update({
            'empleado': empleado,
            'horario_para_hoy': horario_para_hoy,
        })

    return render(request, 'control/asistencias/empleado_dashboard.html', context)


@login_required
def asistencia_events(request):
    """Devuelve eventos de asistencias en formato JSON para el calendario.

    - Si el usuario es administrador y se pasa ?empleado_id=NN filtra por ese empleado.
    - Si no es administrador, devuelve solo las asistencias del empleado asociado al user.
    Cada evento contiene `extendedProps` con información necesaria para el modal.
    """
    user = request.user

    # Base queryset
    if es_administracion(user):
        asistencias = Asistencia.objects.all().select_related('empleado')
        empleado_id = request.GET.get('empleado_id')
        if empleado_id:
            asistencias = asistencias.filter(empleado_id=empleado_id)
    else:
        try:
            empleado = Empleado.objects.get(user=user)
            asistencias = Asistencia.objects.filter(empleado=empleado).select_related('empleado')
        except Empleado.DoesNotExist:
            return JsonResponse([], safe=False)

    eventos = []
    # Mapeo de colores según tipo
    tipo_color = {
        'normal': '#28a745',   # verde
        'retardo': '#ffc107',   # amarillo
        'falta': '#dc3545',     # rojo
        'justificada': '#17a2b8' # cyan/azul
    }

    for a in asistencias:
        color = tipo_color.get(a.tipo, '#6c757d')
        title = a.tipo.title() if not es_administracion(user) else f"{a.empleado.nombre} {a.empleado.apellido} - {a.tipo.title()}"

        eventos.append({
            'id': a.id,
            'title': title,
            'start': a.fecha.isoformat(),
            'allDay': True,
            'backgroundColor': color,
            'borderColor': color,
            'extendedProps': {
                'empleado': str(a.empleado) if a.empleado else None,
                'hora_entrada': a.hora_entrada.strftime('%I:%M %p') if a.hora_entrada else None,
                'hora_salida': a.hora_salida.strftime('%I:%M %p') if a.hora_salida else None,
                'diferencia': a.diferencia if hasattr(a, 'diferencia') else None,
                'tipo': a.tipo,
                'observaciones': a.observaciones,
                'justificante_url': a.justificantes.first().ruta_archivo.url if a.justificantes.exists() and a.justificantes.first().ruta_archivo else None,
                'asistencia_id': a.id
            }
        })

    return JsonResponse(eventos, safe=False)


@login_required
def reporte_asistencias(request):

    if request.GET.get("exportar") == "excel":
        return exportar_asistencias_excel(request)

    """Generar reporte de asistencias (solo administradores)."""
    if not request.user.groups.filter(name='administracion').exists():
        return HttpResponseForbidden('No tienes permiso para ver esta página')

    # Obtener parámetros de filtrado
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    empleado_id = request.GET.get('empleado_id')
    tipo = request.GET.get('tipo')

    # Construir el query base
    asistencias = Asistencia.objects.all().select_related('empleado')

    # Aplicar filtros si existen
    if fecha_inicio:
        asistencias = asistencias.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        asistencias = asistencias.filter(fecha__lte=fecha_fin)
    if empleado_id:
        asistencias = asistencias.filter(empleado_id=empleado_id)
    if tipo:
        asistencias = asistencias.filter(tipo=tipo)

    # Obtener lista de empleados para el filtro
    empleados = Empleado.objects.filter(estado='activo').order_by('nombre')

    return render(request, 'control/asistencias/reporte.html', {
        'asistencias': asistencias,
        'empleados': empleados,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'empleado_id': empleado_id,
        'tipo': tipo
    })


@login_required
def subir_justificante(request, asistencia_id):
    """Subir justificante PDF para un retardo."""
    asistencia = get_object_or_404(Asistencia, pk=asistencia_id)
    user = request.user
    
    # Validar que el empleado solo pueda subir justificantes para sus propias asistencias
    try:
        empleado = Empleado.objects.get(user=user)
        if asistencia.empleado != empleado:
            return HttpResponseForbidden('No tienes permiso para modificar este registro')
    except Empleado.DoesNotExist:
        return HttpResponseForbidden('No tienes acceso a esta página')
    
    # Validar que solo se puedan subir justificantes para retardos
    if asistencia.tipo != 'retardo':
        messages.error(request, 'Solo puedes subir justificantes para asistencias con retardo.')
        return redirect('control:empleado_dashboard')

    if request.method == 'POST':
        form = JustificanteRetardoForm(request.POST, request.FILES)
        if form.is_valid():
            justificante = form.save(commit=False)
            justificante.empleado = asistencia.empleado
            justificante.asistencia = asistencia
            justificante.save()
            messages.success(request, 'Justificante subido exitosamente y está pendiente de validación.')
            return redirect('control:empleado_dashboard')
        else:
            messages.error(request, 'Error al subir el justificante. Verifique el archivo.')
    else:
        form = JustificanteRetardoForm()

    return render(request, 'control/asistencias/subir_justificante.html', {
        'form': form,
        'asistencia': asistencia
    })


@login_required
def validar_justificantes(request):
    """Panel para que el administrador valide justificantes pendientes."""
    if not es_administracion(request.user):
        return HttpResponseForbidden('No tienes permiso para acceder a esta página')
    
    try:
        # Obtener filtros
        estado_filter = request.GET.get('estado', 'pendiente')
        
        # Obtener justificantes
        justificantes = Justificante.objects.select_related('empleado', 'asistencia').order_by('-fecha_envio')
        
        if estado_filter:
            justificantes = justificantes.filter(estado=estado_filter)
        
        # Contar por estado
        stats = {
            'pendiente': Justificante.objects.filter(estado='pendiente').count(),
            'aprobado': Justificante.objects.filter(estado='aprobado').count(),
            'rechazado': Justificante.objects.filter(estado='rechazado').count(),
        }
        
        return render(request, 'control/admin/validar_justificantes.html', {
            'justificantes': justificantes,
            'estado_filter': estado_filter,
            'stats': stats
        })
    except Exception as e:
        logger.error(f"Error en validar_justificantes: {str(e)}")
        messages.error(request, 'La tabla de justificantes aún no existe. Ejecuta las migraciones: python manage.py migrate')
        return redirect('control:admin_dashboard')


@login_required
def aprobar_justificante(request, justificante_id):
    """Aprobar un justificante."""
    if not es_administracion(request.user):
        return HttpResponseForbidden('No tienes permiso para realizar esta acción')
    
    justificante = get_object_or_404(Justificante, pk=justificante_id)
    
    if request.method == 'POST':
        observacion = request.POST.get('observacion', '')
        justificante.estado = 'aprobado'
        justificante.observacion = observacion or 'Aprobado por administrador'
        justificante.save()
        messages.success(request, f'Justificante de {justificante.empleado.nombre} aprobado exitosamente.')
        return redirect('control:validar_justificantes')
    
    return render(request, 'control/admin/detalle_justificante.html', {
        'justificante': justificante,
        'accion': 'aprobar'
    })


@login_required
def rechazar_justificante(request, justificante_id):
    """Rechazar un justificante."""
    if not es_administracion(request.user):
        return HttpResponseForbidden('No tienes permiso para realizar esta acción')
    
    justificante = get_object_or_404(Justificante, pk=justificante_id)
    
    if request.method == 'POST':
        observacion = request.POST.get('observacion', 'Rechazado por administrador')
        justificante.estado = 'rechazado'
        justificante.observacion = observacion
        justificante.save()
        messages.warning(request, f'Justificante de {justificante.empleado.nombre} rechazado.')
        return redirect('control:validar_justificantes')
    
    return render(request, 'control/admin/detalle_justificante.html', {
        'justificante': justificante,
        'accion': 'rechazar'
    })


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from .models import Asistencia, Empleado

@login_required
def exportar_asistencias_excel(request):

    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    empleado_id = request.GET.get("empleado")
    tipo = request.GET.get("tipo")

    asistencias = Asistencia.objects.select_related("empleado").all()

    if fecha_inicio:
        asistencias = asistencias.filter(fecha__gte=fecha_inicio)

    if fecha_fin:
        asistencias = asistencias.filter(fecha__lte=fecha_fin)

    if empleado_id and empleado_id != "todos":
        asistencias = asistencias.filter(empleado_id=empleado_id)

    if tipo and tipo != "todos":
        asistencias = asistencias.filter(tipo=tipo)

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencias"

    ws.merge_cells("A1:F1")
    titulo = ws["A1"]
    titulo.value = "Reporte de Asistencias"
    titulo.font = Font(size=16, bold=True)
    titulo.alignment = Alignment(horizontal="center")

    encabezados = ["Fecha", "Empleado", "Entrada", "Salida", "Tipo", "Observaciones"]
    ws.append(encabezados)

    header_fill = PatternFill(start_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center")

    for col in range(1, len(encabezados) + 1):
        c = ws.cell(row=2, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_alignment

    for a in asistencias:
        ws.append([
            a.fecha.strftime("%d/%m/%Y"),
            str(a.empleado),
            a.hora_entrada.strftime("%H:%M:%S") if a.hora_entrada else "-",
            a.hora_salida.strftime("%H:%M:%S") if a.hora_salida else "-",
            a.get_tipo_display(),
            a.observaciones or "-"
        ])

    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    max_row = ws.max_row
    max_col = ws.max_column

    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value)
            if val:
                max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max_len + 6

    ws.auto_filter.ref = f"A2:{get_column_letter(max_col)}{max_row}"

    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="reporte_asistencias.xlsx"'
    wb.save(response)
    return response
